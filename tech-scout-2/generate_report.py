"""Report generator for horizon scan results.

Reads scan_config.json, scan_results.json, and scan_report.md from the .tmp
directory, then generates HTML and/or Excel output files in the reports/ folder.

Usage:
    python generate_report.py <path_to_scan_config.json>
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path


def _safe_filename(technology: str) -> str:
    clean = re.sub(r"[^\w\s-]", "", technology.lower())
    clean = re.sub(r"[\s]+", "_", clean.strip())
    return clean


def _parse_companies_table(report_md: str) -> list[dict]:
    """Extract rows from the markdown companies table in the report."""
    rows = []
    in_table = False
    for line in report_md.splitlines():
        line = line.strip()
        if line.startswith("| **Organization**") or line.startswith("| Organization"):
            in_table = True
            continue
        if in_table and line.startswith("|---"):
            continue
        if in_table and line.startswith("|") and line.endswith("|"):
            parts = [p.strip() for p in line.strip("|").split("|")]
            if len(parts) >= 5:
                rows.append({
                    "Organization": parts[0],
                    "Type": parts[1],
                    "Status": parts[2],
                    "Evidence": parts[3],
                    "Source": parts[4],
                })
        elif in_table and not line.startswith("|"):
            in_table = False
    return rows


def generate_html(technology: str, report_md: str, output_path: Path) -> None:
    """Convert the markdown report to a styled HTML file."""
    import html as html_lib

    def md_to_html(text: str) -> str:
        lines = text.splitlines()
        out = []
        in_table = False
        i = 0
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()

            if stripped.startswith("## "):
                out.append(f"<h2>{html_lib.escape(stripped[3:])}</h2>")
            elif stripped.startswith("### "):
                out.append(f"<h3>{html_lib.escape(stripped[4:])}</h3>")
            elif stripped.startswith("- "):
                out.append(f"<li>{html_lib.escape(stripped[2:])}</li>")
            elif stripped.startswith("> "):
                out.append(f"<p class='footnote'>{html_lib.escape(stripped[2:])}</p>")
            elif stripped.startswith("| **Organization**") or stripped.startswith("| Organization"):
                in_table = True
                out.append("<table>")
                headers = [h.strip().strip("*") for h in stripped.strip("|").split("|")]
                out.append("<thead><tr>" + "".join(f"<th>{html_lib.escape(h)}</th>" for h in headers) + "</tr></thead>")
                out.append("<tbody>")
            elif in_table and stripped.startswith("|---"):
                pass
            elif in_table and stripped.startswith("|") and stripped.endswith("|"):
                cells = [c.strip() for c in stripped.strip("|").split("|")]
                out.append("<tr>" + "".join(f"<td>{html_lib.escape(c)}</td>" for c in cells) + "</tr>")
            elif in_table and not stripped.startswith("|"):
                out.append("</tbody></table>")
                in_table = False
                if stripped:
                    out.append(f"<p>{html_lib.escape(stripped)}</p>")
            elif stripped == "---":
                out.append("<hr>")
            elif stripped:
                out.append(f"<p>{html_lib.escape(stripped)}</p>")
            i += 1
        if in_table:
            out.append("</tbody></table>")
        return "\n".join(out)

    body = md_to_html(report_md)
    timestamp = datetime.now().strftime("%B %d, %Y")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{html_lib.escape(technology.title())} -- Horizon Scan</title>
  <style>
    body {{ font-family: Georgia, serif; max-width: 960px; margin: 40px auto; padding: 0 24px; color: #1a1a1a; line-height: 1.7; }}
    h2 {{ font-size: 1.6rem; border-bottom: 2px solid #1a1a1a; padding-bottom: 6px; margin-top: 40px; }}
    h3 {{ font-size: 1.15rem; margin-top: 28px; color: #333; }}
    table {{ border-collapse: collapse; width: 100%; margin: 16px 0; font-family: Arial, sans-serif; font-size: 0.9rem; }}
    th {{ background: #1a1a1a; color: #fff; padding: 10px 12px; text-align: left; }}
    td {{ padding: 9px 12px; border-bottom: 1px solid #ddd; vertical-align: top; }}
    tr:nth-child(even) td {{ background: #f9f9f9; }}
    li {{ margin: 4px 0; }}
    hr {{ border: none; border-top: 1px solid #ccc; margin: 32px 0; }}
    .footnote {{ font-style: italic; color: #666; font-size: 0.9rem; }}
    .meta {{ color: #888; font-size: 0.85rem; margin-bottom: 32px; }}
  </style>
</head>
<body>
  <p class="meta">Generated {timestamp} &mdash; BCore Horizon Scan</p>
  {body}
</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")


def generate_excel(technology: str, report_md: str, results: dict, output_path: Path) -> None:
    """Write a structured Excel workbook with summary + companies + raw data sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 70

    ws_summary["A1"] = "Technology"
    ws_summary["B1"] = technology.title()
    ws_summary["A1"].font = header_font

    ws_summary["A2"] = "Context"
    ws_summary["B2"] = results.get("context", "")
    ws_summary["A3"] = "Focus"
    ws_summary["B3"] = results.get("focus", "")
    ws_summary["A4"] = "Geography"
    ws_summary["B4"] = results.get("geography", "")
    ws_summary["A5"] = "Generated"
    ws_summary["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws_summary["A7"] = "Source Counts"
    ws_summary["A7"].font = header_font
    counts = results.get("counts", {})
    row = 8
    for source, count in counts.items():
        ws_summary[f"A{row}"] = source.upper()
        ws_summary[f"B{row}"] = count
        row += 1

    ws_summary[f"A{row+1}"] = "Subtopics Searched"
    ws_summary[f"A{row+1}"].font = header_font
    for j, subtopic in enumerate(results.get("subtopics_searched", []), start=row + 2):
        ws_summary[f"B{j}"] = subtopic

    # --- Sheet 2: Companies ---
    ws_companies = wb.create_sheet("Companies")
    company_headers = ["Organization", "Type", "Status", "Evidence", "Source"]
    for col, header in enumerate(company_headers, start=1):
        cell = ws_companies.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    col_widths = [30, 18, 14, 60, 12]
    for col, width in enumerate(col_widths, start=1):
        ws_companies.column_dimensions[get_column_letter(col)].width = width

    companies = _parse_companies_table(report_md)
    for r, company in enumerate(companies, start=2):
        for col, key in enumerate(company_headers, start=1):
            cell = ws_companies.cell(row=r, column=col, value=company.get(key, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Sheet 3+: Raw Results per source ---
    raw_results = results.get("results", {})
    for source_name, items in raw_results.items():
        if not items:
            continue
        ws = wb.create_sheet(source_name.upper())
        keys = list(items[0].keys())
        for col, key in enumerate(keys, start=1):
            cell = ws.cell(row=1, column=col, value=key.title())
            cell.font = header_font_white
            cell.fill = header_fill
        for r, item in enumerate(items, start=2):
            for col, key in enumerate(keys, start=1):
                val = item.get(key, "")
                if isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                ws.cell(row=r, column=col, value=str(val) if val else "")
        for col in range(1, len(keys) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_report.py <scan_config.json>")
        sys.exit(1)

    config_path = Path(sys.argv[1])
    tmp_dir = config_path.parent

    config = json.loads(config_path.read_text(encoding="utf-8"))
    technology = config["technology"]
    output_format = config.get("output_format", "none").lower()

    if output_format == "none":
        print("No file output requested.")
        return

    results_path = tmp_dir / "scan_results.json"
    report_path = tmp_dir / "scan_report.md"

    if not results_path.exists():
        print(f"ERROR: scan_results.json not found at {results_path}")
        sys.exit(1)
    if not report_path.exists():
        print(f"ERROR: scan_report.md not found at {report_path}")
        sys.exit(1)

    results = json.loads(results_path.read_text(encoding="utf-8"))
    report_md = report_path.read_text(encoding="utf-8")

    reports_dir = Path(__file__).parent / "reports"
    reports_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_name = f"{_safe_filename(technology)}_{timestamp}"

    if output_format in ("html", "both"):
        html_path = reports_dir / f"{base_name}.html"
        generate_html(technology, report_md, html_path)
        print(f"HTML report saved to: {html_path}")

    if output_format in ("excel", "both"):
        xlsx_path = reports_dir / f"{base_name}.xlsx"
        generate_excel(technology, report_md, results, xlsx_path)
        print(f"Excel report saved to: {xlsx_path}")


if __name__ == "__main__":
    main()
