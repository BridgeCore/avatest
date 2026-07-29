#!/usr/bin/env python3
"""
Creates input/sample_gm_data.xlsx — a fictitious Revenue/Gross Margin workbook
with monthly "GM Report" actuals sheets, an "AOP" plan sheet, and monthly
"Compare" sheets. All names and numbers are entirely fictitious.

Divisions: MS1, MS2, MS3, IS1, BL1 (2-3 projects each).
Months: Jan26, Feb26, Mar26 — deliberately using two different GM-Report
sheet-naming conventions ("Jan26" and "Feb26 GM Report") to exercise the
loader's alias/name-detection logic.
"""

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

random.seed(42)

ROOT = Path(__file__).parent

DIVISIONS = ["MS1", "MS2", "MS3", "IS1", "BL1"]

PROJECTS = {
    "MS1": ["Orion Sustainment", "Falcon Modernization"],
    "MS2": ["Trident Upgrade", "Nimbus Integration"],
    "MS3": ["Ember Fuels R&D", "Solara Propellant"],
    "IS1": ["Sentinel Analytics", "Beacon C2"],
    "BL1": ["Forge Labs Support"],
}

PROJECT_CODES = {}
_code_n = 1000
for div, names in PROJECTS.items():
    for name in names:
        PROJECT_CODES[name] = f"{div}-{_code_n}"
        _code_n += 1

MONTHS = [("Jan26", "Jan26"), ("Feb26", "Feb26 GM Report"), ("Mar26", "Mar26 GM Report")]

# Base monthly revenue/GM% per project (revenue drifts +/-8% month to month)
BASE = {
    "Orion Sustainment":     (420_000, 0.34),
    "Falcon Modernization":  (310_000, 0.29),
    "Trident Upgrade":       (505_000, 0.31),
    "Nimbus Integration":    (275_000, 0.27),
    "Ember Fuels R&D":       (190_000, 0.22),
    "Solara Propellant":     (160_000, 0.25),
    "Sentinel Analytics":    (380_000, 0.38),
    "Beacon C2":             (295_000, 0.33),
    "Forge Labs Support":    (120_000, 0.30),
}

# Division-level AOP: monthly revenue & GP targets (Jan-Dec), fictitious
AOP_REVENUE = {
    "MS1": [730_000] * 12,
    "MS2": [780_000] * 12,
    "MS3": [360_000] * 12,
    "IS1": [670_000] * 12,
    "BL1": [125_000] * 12,
}
AOP_GP_PCT = {"MS1": 0.32, "MS2": 0.30, "MS3": 0.24, "IS1": 0.36, "BL1": 0.29}


def _drifted(base: float, month_idx: int) -> float:
    drift = 1.0 + random.uniform(-0.08, 0.08) + month_idx * 0.015
    return round(base * drift, 2)


def build_gm_report_sheet(wb: Workbook, sheet_name: str, month_idx: int):
    ws = wb.create_sheet(sheet_name)
    headers = ["AOP Legend", "Org", "Project Code", "Revenue", "GrossMargin",
               "GrossMarginPercentage", "TotalDirectCost"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for div, names in PROJECTS.items():
        for name in names:
            base_rev, base_gmp = BASE[name]
            rev = _drifted(base_rev, month_idx)
            gmp = round(base_gmp + random.uniform(-0.02, 0.02), 4)
            gm = round(rev * gmp, 2)
            tdc = round(rev - gm, 2)
            ws.append([name, div, PROJECT_CODES[name], rev, gm, gmp, tdc])


def build_aop_sheet(wb: Workbook):
    ws = wb.create_sheet("AOP")
    # Column C (index 3, 1-based) = division; columns D-O (4-15) = Jan-Dec
    ws.cell(row=1, column=3, value="Revenue Plan")
    for m in range(12):
        ws.cell(row=1, column=4 + m, value=f"M{m+1}")

    # Revenue rows: Excel rows 3-7
    for i, div in enumerate(DIVISIONS):
        r = 3 + i
        ws.cell(row=r, column=3, value=div)
        for m in range(12):
            ws.cell(row=r, column=4 + m, value=AOP_REVENUE[div][m])

    ws.cell(row=11, column=3, value="Gross Profit Plan")
    # GP rows: Excel rows 12-16
    for i, div in enumerate(DIVISIONS):
        r = 12 + i
        ws.cell(row=r, column=3, value=div)
        for m in range(12):
            gp = round(AOP_REVENUE[div][m] * AOP_GP_PCT[div], 2)
            ws.cell(row=r, column=4 + m, value=gp)


def build_compare_sheet(wb: Workbook, month_label: str, month_idx: int):
    ws = wb.create_sheet(f"{month_label} Compare")
    headers = ["Program", "Portfolio", "Actuals", "AOP", "Var", "GP Actual", "AOP GM%", "GP AOP"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    portfolio_leads = {"MS1": "Dana Reyes", "MS2": "Marcus Webb", "MS3": "Priya Anand",
                        "IS1": "Sofia Chen", "BL1": "Owen Baptiste"}

    for div, names in PROJECTS.items():
        n_projects = len(names)
        div_aop_rev_month = AOP_REVENUE[div][month_idx] / n_projects
        div_gp_pct = AOP_GP_PCT[div]
        for name in names:
            base_rev, base_gmp = BASE[name]
            actual = _drifted(base_rev, month_idx)
            aop = round(div_aop_rev_month, 2)
            var = round(actual - aop, 2)
            gp_actual = round(actual * (base_gmp + random.uniform(-0.02, 0.02)), 2)
            gp_aop = round(aop * div_gp_pct, 2)
            ws.append([name, portfolio_leads[div], actual, aop, var, gp_actual, div_gp_pct, gp_aop])


def main():
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    for month_idx, (short, sheet_name) in enumerate(MONTHS):
        build_gm_report_sheet(wb, sheet_name, month_idx)
        build_compare_sheet(wb, short, month_idx)

    build_aop_sheet(wb)

    out_dir = ROOT / "input"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "sample_gm_data.xlsx"
    wb.save(out_path)
    print(f"Wrote {out_path} ({len(MONTHS)} month(s), {sum(len(v) for v in PROJECTS.values())} projects)")


if __name__ == "__main__":
    main()
