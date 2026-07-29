#!/usr/bin/env python3
"""
Creates input/sample_data.xlsx with all 8 required sheets.
All names and numbers are entirely fictitious.

Employee patterns:
  Emp A: persistent low billable util, unexplained (critical)
  Emp B: persistent low billable util, explained by high PTO (warning)
  Emp C: persistent high B&P hours (warning)
  Emp D: persistent high non-billable project ratio (warning)
  Emp E: G&A crowding out billable util (warning)
  Emp F: partial period — started 2026-02-10
  Emp G: PT employee
  Emp H: LWOP across multiple periods, explains low utilization
  Emp I: corporate role — 90% G&A, zero billable (excluded from trends)
  Emp J: in Time Details, not in PTO Balances (data quality)
  Emp K: in PTO Balances, not in Time Details (data quality)
  Emp L-O: clean, on-target employees
"""

import sys
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

PERIODS = [
    (date(2026, 1,  1), date(2026, 1, 15)),
    (date(2026, 1, 16), date(2026, 1, 31)),
    (date(2026, 2,  1), date(2026, 2, 15)),
    (date(2026, 2, 16), date(2026, 2, 28)),
    (date(2026, 3,  1), date(2026, 3, 15)),
    (date(2026, 3, 16), date(2026, 3, 31)),
    (date(2026, 4,  1), date(2026, 4, 15)),
    (date(2026, 4, 16), date(2026, 4, 30)),
    (date(2026, 5,  1), date(2026, 5, 15)),
    (date(2026, 5, 16), date(2026, 5, 31)),
]

def networkdays(start: date, end: date) -> int:
    return int(np.busday_count(start, end + timedelta(days=1)))

NETWORKDAYS = [networkdays(s, e) for s, e in PERIODS]
NET_HOURS   = [nd * 8 for nd in NETWORKDAYS]

# Employees: (name, org, division, portfolio_lead, contract_type)
EMPLOYEES = [
    ("Alice Abernathy",   "Ops Group",    "BL",   "Greg",  "LH"),
    ("Brian Bosworth",    "Ops Group",    "BL",   "Greg",  "FP"),
    ("Carla Cortez",      "Intel Systems","IS",   "Mason", "LH"),
    ("David Drummond",    "Intel Systems","IS",   "Mason", "LH"),
    ("Elena Everett",     "Intel Systems","IS",   "Mason", "FP"),
    ("Frank Fujimoto",    "Mission Sys",  "MS1",  "Mike",  "FP"),
    ("Grace Gomez",       "Ops Group",    "BL",   "Greg",  "LH"),
    ("Henry Harrington",  "Mission Sys",  "MS1",  "Mike",  "LH"),
    ("Ivan Ingram",       "Corp HQ",      "Corp", "Troy",  ""),
    ("Julia Jones",       "Ops Group",    "BL",   "Roger", "LH"),
    ("Karen Kim",         "Intel Systems","IS",   "Mason", "FP"),
    ("Liam Lawson",       "Mission Sys",  "MS2",  "Mike",  "LH"),
    ("Mia Morales",       "Mission Sys",  "MS2",  "Roger", "FP"),
    ("Nate Nelson",       "Corp HQ",      "BL",   "Troy",  "LH"),
    ("Olivia Owens",      "Intel Systems","IS",   "Mason", "LH"),
]

# Names convenience
EMP = {e[0]: e for e in EMPLOYEES}

# Subgroup labels
SG_BILLABLE    = "ProjectBillable"
SG_NONBILLABLE = "ProjectNonBillable"
SG_BP          = "B&P"
SG_GA          = "G&A"
SG_IRD         = "IR&D"
SG_OVERHEAD    = "Overhead"
SG_PTO         = "PTO"
SG_HOLIDAY     = "Holiday"
SG_LWOP        = "LWOP"
SG_OTHER       = "Other"

PC_RT   = "RT"
PC_PTO  = "PTO"
PC_HOL  = "HOL"
PC_LWOP = "LWOP"
PC_UL   = "UL"   # fringe rows — will be excluded

PROJECT_CODES = {
    SG_BILLABLE:    ("P-1001", "Billable Task A"),
    SG_NONBILLABLE: ("P-1002", "NonBillable Task"),
    SG_BP:          ("P-9001", "B&P General"),
    SG_GA:          ("P-9002", "G&A Admin"),
    SG_IRD:         ("P-9003", "IR&D Research"),
    SG_OVERHEAD:    ("P-9004", "Overhead"),
    SG_PTO:         ("P-8001", "PTO Leave"),
    SG_HOLIDAY:     ("P-8002", "Holiday"),
    SG_LWOP:        ("P-8003", "LWOP"),
    SG_OTHER:       ("P-8004", "Other"),
}

def _paycode(sg: str) -> str:
    return {SG_PTO: PC_PTO, SG_HOLIDAY: PC_HOL, SG_LWOP: PC_LWOP}.get(sg, PC_RT)


def _period_mid(period_idx: int) -> date:
    s, e = PERIODS[period_idx]
    return s + timedelta(days=(e - s).days // 2)


# ─────────────────────────────────────────────────────────────────────────────
# Hour schedule builder
# ─────────────────────────────────────────────────────────────────────────────

def make_rows(name: str, period_idx: int, hours_by_sg: dict) -> list[dict]:
    """Build Export rows for one employee in one period."""
    etype = EMP.get(name)
    if etype is None:
        return []
    _, org, div, lead, _ = etype
    dt = _period_mid(period_idx)

    rows = []
    for sg, hrs in hours_by_sg.items():
        if hrs == 0:
            continue
        proj_code, proj_title = PROJECT_CODES[sg]
        rows.append({
            "PersonOrganization": org,
            "Person": name,
            "ProjectOrganization": org,
            "ProjectCode": proj_code,
            "ProjectTitle": proj_title,
            "ProjectOwningOrg": org,
            "TaskNumber": "001",
            "Task": "Default",
            "LaborCategory": "Sr Analyst",
            "Location": "Remote",
            "ProjectType": "CPFF" if sg in (SG_BILLABLE, SG_NONBILLABLE) else "Indirect",
            "PayCode": _paycode(sg),
            "Reference": "",
            "Date": dt,
            "ADJPostedDate": dt,
            "FinancialPostedDate": dt,
            "Hours": hrs,
            "PersonDivision": div,
            "ProjectType-PayCode": f"CPFF-{_paycode(sg)}",
            "ProjectGroup": ("Productive" if sg in (SG_BILLABLE,) else
                             "Project" if sg in (SG_NONBILLABLE, SG_BP, SG_GA, SG_IRD, SG_OVERHEAD) else
                             "TimeOff"),
            "ProjectSubGroup": sg,
        })
    return rows


def make_ul_rows(name: str, period_idx: int, hrs: float = 2.0) -> list[dict]:
    """Fringe rows with PayCode=UL — must be excluded from analysis."""
    etype = EMP.get(name)
    if etype is None:
        return []
    _, org, div, _, _ = etype
    dt = _period_mid(period_idx)
    return [{
        "PersonOrganization": org, "Person": name,
        "ProjectOrganization": org, "ProjectCode": "P-UL01",
        "ProjectTitle": "Fringe Benefits", "ProjectOwningOrg": org,
        "TaskNumber": "001", "Task": "Fringe", "LaborCategory": "N/A",
        "Location": "N/A", "ProjectType": "Indirect", "PayCode": PC_UL,
        "Reference": "", "Date": dt, "ADJPostedDate": dt, "FinancialPostedDate": dt,
        "Hours": hrs, "PersonDivision": div,
        "ProjectType-PayCode": f"Indirect-{PC_UL}",
        "ProjectGroup": "Project", "ProjectSubGroup": SG_GA,
    }]


# ─────────────────────────────────────────────────────────────────────────────
# Build Export rows per employee pattern
# ─────────────────────────────────────────────────────────────────────────────

def build_export_rows() -> list[dict]:
    rows = []

    for pi in range(10):
        nh = NET_HOURS[pi]

        # ── Emp A: Alice Abernathy — persistent low billable, unexplained (CRITICAL) ──
        rows += make_rows("Alice Abernathy", pi, {SG_BILLABLE: nh * 0.30, SG_OVERHEAD: nh * 0.10, SG_HOLIDAY: min(8, nh * 0.05)})

        # ── Emp B: Brian Bosworth — low billable, explained by high PTO (WARNING after downgrade) ──
        if pi >= 2:  # start after 2 clean periods
            rows += make_rows("Brian Bosworth", pi, {SG_BILLABLE: nh * 0.40, SG_PTO: 24.0, SG_OVERHEAD: nh * 0.05})
        else:
            rows += make_rows("Brian Bosworth", pi, {SG_BILLABLE: nh * 0.80, SG_HOLIDAY: 8.0})

        # ── Emp C: Carla Cortez — persistent high B&P ──
        rows += make_rows("Carla Cortez", pi, {SG_BILLABLE: nh * 0.50, SG_BP: nh * 0.30, SG_HOLIDAY: 8.0})

        # ── Emp D: David Drummond — persistent high non-billable ratio ──
        rows += make_rows("David Drummond", pi, {SG_BILLABLE: nh * 0.40, SG_NONBILLABLE: nh * 0.35, SG_HOLIDAY: 8.0})

        # ── Emp E: Elena Everett — G&A crowding ──
        rows += make_rows("Elena Everett", pi, {SG_BILLABLE: nh * 0.20, SG_GA: nh * 0.55, SG_HOLIDAY: 8.0})

        # ── Emp F: Frank Fujimoto — partial period (started 2026-02-10) ──
        if pi >= 2:  # active from period 3 onwards
            rows += make_rows("Frank Fujimoto", pi, {SG_BILLABLE: nh * 0.80, SG_HOLIDAY: 8.0})
        # no rows in periods 1-2 (before start date)

        # ── Emp G: Grace Gomez — PT employee ──
        pt_hrs = 60.0  # variable schedule, roughly half-time
        rows += make_rows("Grace Gomez", pi, {SG_BILLABLE: pt_hrs * 0.72, SG_BP: pt_hrs * 0.10, SG_HOLIDAY: 4.0})

        # ── Emp H: Henry Harrington — LWOP explains low utilization ──
        if pi >= 1:  # LWOP from period 2 onwards
            rows += make_rows("Henry Harrington", pi, {SG_BILLABLE: nh * 0.20, SG_LWOP: 48.0, SG_HOLIDAY: 8.0})
        else:
            rows += make_rows("Henry Harrington", pi, {SG_BILLABLE: nh * 0.80, SG_HOLIDAY: 8.0})

        # ── Emp I: Ivan Ingram — corporate role, 90% G&A, zero billable ──
        rows += make_rows("Ivan Ingram", pi, {SG_GA: nh * 0.85, SG_OVERHEAD: nh * 0.05, SG_HOLIDAY: 8.0})

        # ── Emp J: Julia Jones — in Time Details, not in PTO Balances (data quality) ──
        rows += make_rows("Julia Jones", pi, {SG_BILLABLE: nh * 0.78, SG_HOLIDAY: 8.0})

        # Emp K (Karen Kim) — in PTO Balances, no Time Details; no Export rows

        # ── Emp L-O: clean on-target employees ──
        rows += make_rows("Liam Lawson",    pi, {SG_BILLABLE: nh * 0.78, SG_BP: nh * 0.05, SG_HOLIDAY: 8.0})
        rows += make_rows("Mia Morales",    pi, {SG_BILLABLE: nh * 0.82, SG_GA: nh * 0.03, SG_HOLIDAY: 8.0})
        rows += make_rows("Nate Nelson",    pi, {SG_BILLABLE: nh * 0.76, SG_OVERHEAD: nh * 0.04, SG_HOLIDAY: 8.0})
        rows += make_rows("Olivia Owens",   pi, {SG_BILLABLE: nh * 0.80, SG_BP: nh * 0.03, SG_HOLIDAY: 8.0})

        # Add UL fringe rows (should be excluded) — sprinkle across a few employees
        if pi % 3 == 0:
            rows += make_ul_rows("Alice Abernathy", pi)
            rows += make_ul_rows("Liam Lawson", pi)
            rows += make_ul_rows("Mia Morales", pi)

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Sheet writers
# ─────────────────────────────────────────────────────────────────────────────

def write_billable_ute(ws):
    ws.title = "Billable Ute"
    ws["A1"] = "Period Start Date"
    ws["B1"] = PERIODS[0][0]
    ws["A2"] = "Period End Date"
    ws["B2"] = PERIODS[-1][1]
    ws["B1"].number_format = "YYYY-MM-DD"
    ws["B2"].number_format = "YYYY-MM-DD"

    headers = [
        "PersonOrganization", "Person", "Division", "Portfolio Lead",
        "Billable Util", "Direct Util", "Total Hrs",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h).font = Font(bold=True)

    for ri, (name, org, div, lead, _) in enumerate(EMPLOYEES, 4):
        ws.cell(row=ri, column=1, value=org)
        ws.cell(row=ri, column=2, value=name)
        ws.cell(row=ri, column=3, value=div)
        ws.cell(row=ri, column=4, value=lead)
        ws.cell(row=ri, column=5, value=0.75)  # placeholder
        ws.cell(row=ri, column=6, value=0.80)
        ws.cell(row=ri, column=7, value=120.0)


def write_export(ws, rows: list[dict]):
    ws.title = "Export"
    HEADERS = [
        "PersonOrganization", "Person", "ProjectOrganization", "ProjectCode",
        "ProjectTitle", "ProjectOwningOrg", "TaskNumber", "Task", "LaborCategory",
        "Location", "ProjectType", "PayCode", "Reference", "Date",
        "ADJPostedDate", "FinancialPostedDate", "Hours", "PersonDivision",
        "ProjectType-PayCode", "ProjectGroup", "ProjectSubGroup",
    ]
    for ci, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(HEADERS, 1):
            val = row.get(h, "")
            ws.cell(row=ri, column=ci, value=val)


def write_lookups(ws):
    ws.title = "Lookups"

    # ── Section A: Period calendar starting at column G (col 7) ────────────────
    ws.cell(row=1, column=1, value="PT Employees")
    ws.cell(row=1, column=1).font = Font(bold=True)

    # Section B: PT employees
    pt_emps = ["Grace Gomez"]
    for ri, name in enumerate(pt_emps, 2):
        ws.cell(row=ri, column=1, value=name)

    # Section C: Personnel lookup — PersonOrganization, Division, Portfolio Lead, PrimaryContractType
    lookup_header_col = 3
    ws.cell(row=1, column=lookup_header_col,     value="PersonOrganization").font = Font(bold=True)
    ws.cell(row=1, column=lookup_header_col + 1, value="Division").font = Font(bold=True)
    ws.cell(row=1, column=lookup_header_col + 2, value="Portfolio Lead").font = Font(bold=True)
    ws.cell(row=1, column=lookup_header_col + 3, value="PrimaryContractType").font = Font(bold=True)

    for ri, (name, org, div, lead, ct) in enumerate(EMPLOYEES, 2):
        ws.cell(row=ri, column=lookup_header_col,     value=name)
        ws.cell(row=ri, column=lookup_header_col + 1, value=div)
        ws.cell(row=ri, column=lookup_header_col + 2, value=lead)
        ws.cell(row=ri, column=lookup_header_col + 3, value=ct)

    # Section A: Period calendar at column G (index 7)
    START_COL = 7  # column G
    ws.cell(row=1, column=START_COL - 1, value="Period Calendar →").font = Font(bold=True)

    for i, (start, end) in enumerate(PERIODS):
        col = START_COL + i
        nd  = NETWORKDAYS[i]
        nh  = NET_HOURS[i]
        ws.cell(row=1, column=col, value=start)
        ws.cell(row=1, column=col).number_format = "YYYY-MM-DD"
        ws.cell(row=2, column=col, value=nd)
        ws.cell(row=3, column=col, value=nh)


def write_unique_employees(ws):
    ws.title = "Unique Employees"
    ws.cell(row=1, column=1, value="PersonOrganization").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Person").font = Font(bold=True)

    # Include Karen Kim here (has PTO record, no time entries)
    all_roster = EMPLOYEES + [("Intel Systems", "Karen Kim", "IS", "Mason", "FP")]

    for ri, (name, org, div, lead, ct) in enumerate(all_roster, 2):
        ws.cell(row=ri, column=1, value=org)
        ws.cell(row=ri, column=2, value=name)


def write_first_last(ws):
    ws.title = "First Days Last Days"
    headers = ["PersonOrganization", "Person", "2026 First Day", "2026 Last Day", "Sortable"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

    # Frank Fujimoto started 2026-02-10
    ws.cell(row=2, column=1, value="Mission Sys")
    ws.cell(row=2, column=2, value="Frank Fujimoto")
    ws.cell(row=2, column=3, value=date(2026, 2, 10))
    ws.cell(row=2, column=3).number_format = "YYYY-MM-DD"
    ws.cell(row=2, column=4, value=None)
    ws.cell(row=2, column=5, value="2026-02-10")


def write_pto_balances(ws):
    ws.title = "Current PTO Balances"
    headers = [
        "Accrual Plan", "Person", "Bus Week Hours", "Hire Date",
        "Beginning Balance", "Period Hours Accrued", "Period Hours Total Balance",
        "Period Hours Used", "Period Hours Available",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

    # All employees EXCEPT Julia Jones (she has no PTO record — data quality)
    # Karen Kim IS in PTO balances (she has no time entries — data quality)
    pto_employees = [e for e in EMPLOYEES if e[0] != "Julia Jones"] + \
                    [("Intel Systems", "Karen Kim", "IS", "Mason", "FP")]

    for ri, (name, org, div, lead, _) in enumerate(pto_employees, 2):
        ws.cell(row=ri, column=1, value="Standard")
        ws.cell(row=ri, column=2, value=name)
        ws.cell(row=ri, column=3, value=40.0)
        ws.cell(row=ri, column=4, value=date(2020, 3, 15))
        ws.cell(row=ri, column=4).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=5, value=80.0)
        ws.cell(row=ri, column=6, value=4.0)
        ws.cell(row=ri, column=7, value=84.0)
        ws.cell(row=ri, column=8, value=24.0)
        ws.cell(row=ri, column=9, value=60.0)


def write_portfolio(ws):
    ws.title = "Portfolio Leads Proj Types"
    ws.cell(row=1, column=1, value="Person").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Portfolio Lead").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Project Type").font = Font(bold=True)

    for ri, (name, org, div, lead, ct) in enumerate(EMPLOYEES, 2):
        ws.cell(row=ri, column=1, value=name)
        ws.cell(row=ri, column=2, value=lead)
        ws.cell(row=ri, column=3, value=ct)


def write_discrepancies(ws):
    ws.title = "Discrepancies"

    # Pipeline Summary
    ws.cell(row=1, column=1, value="Pipeline Summary").font = Font(bold=True)
    ws.cell(row=2, column=1, value="Run Date")
    ws.cell(row=2, column=2, value=str(date.today()))
    ws.cell(row=3, column=1, value="Export Row Count")
    ws.cell(row=3, column=2, value="~500 (sample)")
    ws.cell(row=4, column=1, value="Unique Employees")
    ws.cell(row=4, column=2, value=str(len(EMPLOYEES) + 1))  # +1 for Karen

    # Section A: Name Mismatches
    ws.cell(row=6, column=1, value="Name Mismatches").font = Font(bold=True)
    ws.cell(row=7, column=1, value="Name").font = Font(bold=True)
    ws.cell(row=7, column=2, value="Found In").font = Font(bold=True)
    ws.cell(row=7, column=3, value="Missing From").font = Font(bold=True)

    # Julia Jones: in Time Details, not in PTO Balances
    ws.cell(row=8, column=1, value="Julia Jones")
    ws.cell(row=8, column=2, value="Time Details")
    ws.cell(row=8, column=3, value="PTO Balances")

    # Karen Kim: in PTO Balances, not in Time Details
    ws.cell(row=9, column=1, value="Karen Kim")
    ws.cell(row=9, column=2, value="PTO Balances")
    ws.cell(row=9, column=3, value="Time Details")

    # Section B: Missing Lookup Combos
    ws.cell(row=11, column=1, value="Missing Lookup Combos").font = Font(bold=True)
    ws.cell(row=12, column=1, value="ProjectType").font = Font(bold=True)
    ws.cell(row=12, column=2, value="PayCode").font = Font(bold=True)
    ws.cell(row=13, column=1, value="None — all combos are in Lookups.")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    out_path = Path(__file__).parent / "input" / "sample_data.xlsx"
    wb = Workbook()

    # Remove default sheet
    ws_default = wb.active
    wb.remove(ws_default)

    # Sheet 1: Billable Ute
    write_billable_ute(wb.create_sheet("Billable Ute"))

    # Sheet 2: Export
    export_rows = build_export_rows()
    write_export(wb.create_sheet("Export"), export_rows)

    # Sheet 3: Lookups
    write_lookups(wb.create_sheet("Lookups"))

    # Sheet 4: Unique Employees
    write_unique_employees(wb.create_sheet("Unique Employees"))

    # Sheet 5: First Days Last Days
    write_first_last(wb.create_sheet("First Days Last Days"))

    # Sheet 6: Current PTO Balances
    write_pto_balances(wb.create_sheet("Current PTO Balances"))

    # Sheet 7: Portfolio Leads Proj Types
    write_portfolio(wb.create_sheet("Portfolio Leads Proj Types"))

    # Sheet 8: Discrepancies
    write_discrepancies(wb.create_sheet("Discrepancies"))

    wb.save(out_path)
    ul_count = sum(1 for r in export_rows if r["PayCode"] == "UL")
    print(f"  [OK] Wrote {out_path}")
    print(f"    Export rows: {len(export_rows)} (including {ul_count} UL fringe rows)")
    print(f"    Employees in roster: {len(EMPLOYEES) + 1}")  # +Karen Kim
    print(f"    Periods: {len(PERIODS)}")


if __name__ == "__main__":
    main()
