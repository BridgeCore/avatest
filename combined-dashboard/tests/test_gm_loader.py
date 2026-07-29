import io
from pathlib import Path

import openpyxl
import pytest
import yaml

from src.gm_loader import (
    extract_month,
    is_aop_sheet,
    is_compare_sheet,
    is_gm_sheet,
    load_gm_workbook,
)

ROOT = Path(__file__).parent.parent


@pytest.fixture
def cfg():
    with open(ROOT / "config.yaml", "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet-name classification
# ─────────────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("name", ["Jan26", "Feb 2026 GM Report", "March GM_Report", "AprGMReport"])
def test_is_gm_sheet_true(name):
    assert is_gm_sheet(name)


@pytest.mark.parametrize("name", ["AOP", "Compare", "Jan26 Compare", "Random Sheet"])
def test_is_gm_sheet_false(name):
    assert not is_gm_sheet(name)


def test_is_aop_sheet_excludes_compare():
    assert is_aop_sheet("AOP")
    assert is_aop_sheet("2026 AOP Plan")
    assert not is_aop_sheet("AOP Compare")
    assert not is_aop_sheet("Jan26")


def test_is_compare_sheet():
    assert is_compare_sheet("Jan26 Compare")
    assert is_compare_sheet("compare")
    assert not is_compare_sheet("AOP")


def test_extract_month_parses_two_digit_and_four_digit_years():
    mo = extract_month("Jan26")
    assert mo.idx == 0 and mo.year == 2026 and mo.key == 2026 * 12

    mo2 = extract_month("Feb 2026 GM Report")
    assert mo2.idx == 1 and mo2.year == 2026

    assert extract_month("Not A Month") is None


# ─────────────────────────────────────────────────────────────────────────────
# Workbook parsing — warn-don't-crash behavior
# ─────────────────────────────────────────────────────────────────────────────

def _wb_bytes(build_fn) -> io.BytesIO:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_fn(wb)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_load_gm_workbook_parses_a_well_formed_gm_sheet(cfg):
    def build(wb):
        ws = wb.create_sheet("Jan26")
        ws.append(["AOP Legend", "Org", "Project Code", "Revenue", "GrossMargin",
                    "GrossMarginPercentage", "TotalDirectCost"])
        ws.append(["Orion Sustainment", "MS1", "MS1-1000", 100000, 30000, 0.30, 70000])

    data = load_gm_workbook(_wb_bytes(build), cfg)

    assert len(data.actuals) == 1
    assert data.actuals[0].revenue == 100000
    assert data.actuals[0].division == "MS1"
    assert len(data.months) == 1
    scan = data.sheet_log[0]
    assert scan.type == "gm"
    assert not scan.issues


def test_load_gm_workbook_records_issue_for_missing_columns_without_crashing(cfg):
    def build(wb):
        ws = wb.create_sheet("Jan26")
        ws.append(["AOP Legend", "Org", "Revenue"])  # missing several required columns
        ws.append(["Orion Sustainment", "MS1", 100000])

    data = load_gm_workbook(_wb_bytes(build), cfg)

    assert data.actuals == []
    scan = data.sheet_log[0]
    assert scan.type == "gm"
    assert scan.issues
    assert "Missing required columns" in scan.issues[0]


def test_load_gm_workbook_records_issue_for_unparseable_month(cfg):
    def build(wb):
        ws = wb.create_sheet("GM Report")  # matches is_gm_sheet but has no month token
        ws.append(["AOP Legend", "Org", "Revenue", "GrossMargin", "GrossMarginPercentage", "TotalDirectCost"])
        ws.append(["Orion Sustainment", "MS1", 100000, 30000, 0.30, 70000])

    data = load_gm_workbook(_wb_bytes(build), cfg)

    assert data.actuals == []
    assert "month" in data.sheet_log[0].issues[0].lower()


def test_load_gm_workbook_ignores_rows_outside_known_divisions(cfg):
    def build(wb):
        ws = wb.create_sheet("Jan26")
        ws.append(["AOP Legend", "Org", "Revenue", "GrossMargin", "GrossMarginPercentage", "TotalDirectCost"])
        ws.append(["Some Corp Overhead Row", "CORP", 999999, 0, 0, 999999])  # not a tracked division
        ws.append(["Orion Sustainment", "MS1", 100000, 30000, 0.30, 70000])

    data = load_gm_workbook(_wb_bytes(build), cfg)
    assert len(data.actuals) == 1
    assert data.actuals[0].division == "MS1"


def test_load_gm_workbook_normalizes_division_casing(cfg):
    # Regression: real workbooks have data-entry inconsistencies like "Bl1"
    # instead of "BL1". An exact-case membership check silently dropped these
    # rows (no visible error, just missing revenue) — division values must be
    # uppercased before the known-divisions check, and the scan log should
    # surface that a correction happened.
    def build(wb):
        ws = wb.create_sheet("Jan26")
        ws.append(["AOP Legend", "Org", "Revenue", "GrossMargin", "GrossMarginPercentage", "TotalDirectCost"])
        ws.append(["TO5", "Bl1", 88868.52, 52832.91, 0.594, 36035.61])

    data = load_gm_workbook(_wb_bytes(build), cfg)

    assert len(data.actuals) == 1
    assert data.actuals[0].division == "BL1"
    scan = data.sheet_log[0]
    assert any("non-standard division casing" in i for i in scan.issues)


def test_load_gm_workbook_parses_aop_fixed_cell_layout(cfg):
    def build(wb):
        ws = wb.create_sheet("AOP")
        ws.cell(row=3, column=3, value="MS1")
        for m in range(12):
            ws.cell(row=3, column=4 + m, value=100000 + m * 1000)
        ws.cell(row=12, column=3, value="MS1")
        for m in range(12):
            ws.cell(row=12, column=4 + m, value=30000 + m * 100)

    data = load_gm_workbook(_wb_bytes(build), cfg)
    assert data.has_aop
    assert data.aop["MS1"].revenue[0] == 100000
    assert data.aop["MS1"].gp[0] == 30000
    scan = data.sheet_log[0]
    assert scan.type == "aop"
    assert not scan.issues


def test_load_gm_workbook_aop_sheet_with_no_matching_divisions_reports_issue(cfg):
    def build(wb):
        ws = wb.create_sheet("AOP")
        ws.cell(row=3, column=3, value="Not A Division")
        ws.cell(row=3, column=4, value=100000)

    data = load_gm_workbook(_wb_bytes(build), cfg)
    assert not data.has_aop
    scan = data.sheet_log[0]
    assert scan.issues
    assert "No divisions matched" in scan.issues[0]


def test_load_gm_workbook_a_single_bad_sheet_does_not_abort_the_others(cfg):
    def build(wb):
        good = wb.create_sheet("Jan26")
        good.append(["AOP Legend", "Org", "Revenue", "GrossMargin", "GrossMarginPercentage", "TotalDirectCost"])
        good.append(["Orion Sustainment", "MS1", 100000, 30000, 0.30, 70000])

        bad = wb.create_sheet("Feb26 GM Report")
        bad.append(["Not", "The", "Right", "Columns"])
        bad.append([1, 2, 3, 4])

    data = load_gm_workbook(_wb_bytes(build), cfg)
    assert len(data.actuals) == 1  # Jan26 still parsed
    assert len(data.sheet_log) == 2
    assert data.sheet_log[1].issues  # Feb26 sheet flagged, not fatal


def test_load_gm_workbook_parses_compare_sheet(cfg):
    def build(wb):
        ws = wb.create_sheet("Jan26 Compare")
        ws.append(["Program", "Portfolio", "Actuals", "AOP", "Var", "GP Actual", "AOP GM%", "GP AOP"])
        ws.append(["Orion Sustainment", "Dana Reyes", 100000, 95000, 5000, 30000, 0.30, 28500])

    data = load_gm_workbook(_wb_bytes(build), cfg)
    assert len(data.compare) == 1
    assert data.compare[0].program == "Orion Sustainment"
    assert data.compare[0].var_rev == 5000


def test_load_gm_workbook_attaches_portfolio_lead_from_compare_data(cfg):
    def build(wb):
        gm = wb.create_sheet("Jan26")
        gm.append(["AOP Legend", "Org", "Revenue", "GrossMargin", "GrossMarginPercentage", "TotalDirectCost"])
        gm.append(["Orion Sustainment", "MS1", 100000, 30000, 0.30, 70000])

        cmp_ = wb.create_sheet("Jan26 Compare")
        cmp_.append(["Program", "Portfolio", "Actuals", "AOP", "Var", "GP Actual", "AOP GM%", "GP AOP"])
        cmp_.append(["Orion Sustainment", "Dana Reyes", 100000, 95000, 5000, 30000, 0.30, 28500])

    data = load_gm_workbook(_wb_bytes(build), cfg)
    project = next(iter(data.projects.values()))
    assert project.portfolio_lead == "Dana Reyes"
